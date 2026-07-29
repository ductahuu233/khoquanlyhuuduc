import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = os.path.join(os.getcwd(), "exports")

def generate_excel(export_data: dict) -> str:
    """
    Sinh file Excel Nhật Ký Xuất Kho cho phiếu xuất (so_nhat_ky_xuat_kho_{request_id}.xlsx).
    """
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    request_id = export_data.get("request_id", "000")
    file_path = os.path.join(OUTPUT_DIR, f"so_nhat_ky_xuat_kho_{request_id}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhật Ký Xuất Kho"
    _create_headers(ws)

    stt = 1
    export_date = export_data.get("export_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    requester = export_data.get("requester_name", "")
    destination = export_data.get("destination", "N/A")
    reason = export_data.get("reason", "N/A")
    items = export_data.get("items", [])

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for item in items:
        row_data = [
            stt,
            f"PXK-{request_id}",
            export_date,
            requester,
            destination,
            reason,
            item.get("item_code", ""),
            item.get("name", ""),
            item.get("unit", ""),
            item.get("quantity", 0)
        ]
        ws.append(row_data)
        stt += 1

        current_row = ws.max_row
        for col_num in range(1, 11):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            if col_num in [1, 2, 3, 7, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(file_path)

    # Update master log file `bao_cao_xuat_kho.xlsx` for overall audit
    try:
        master_path = os.path.join(OUTPUT_DIR, "bao_cao_xuat_kho.xlsx")
        if os.path.exists(master_path):
            m_wb = openpyxl.load_workbook(master_path)
            m_ws = m_wb.active
        else:
            m_wb = openpyxl.Workbook()
            m_ws = m_wb.active
            m_ws.title = "Tổng Hợp Xuất Kho"
            _create_headers(m_ws)
        
        for item in items:
            row_data = [
                m_ws.max_row,
                f"PXK-{request_id}",
                export_date,
                requester,
                destination,
                reason,
                item.get("item_code", ""),
                item.get("name", ""),
                item.get("unit", ""),
                item.get("quantity", 0)
            ]
            m_ws.append(row_data)
        m_wb.save(master_path)
    except Exception as e:
        print("Update master excel log error:", e)

    return file_path

def _create_headers(ws):
    headers = [
        "STT", "Mã Phiếu", "Ngày Xuất", "Người Yêu Cầu", "Đơn Vị Nhận (Xuất Đi)", "Lý Do Xuất",
        "Mã Vật Tư", "Tên Vật Tư", "Đơn Vị Tính", "Số Lượng Xuất"
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    for col_num in range(1, 11):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [8, 15, 20, 22, 25, 25, 15, 30, 15, 15]
    for i, col_width in enumerate(col_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = col_width
