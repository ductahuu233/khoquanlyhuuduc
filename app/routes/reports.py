import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import Item, Asset, AssetHistory, AuditSheet, log_audit
from app.schemas import StockReconciliationPayload

router = APIRouter(prefix="/api/reports", tags=["Reports & Stock Reconciliation"])

EXPORTS_DIR = os.path.join(os.getcwd(), "exports")

@router.get("/stock-card/{item_id}")
def get_stock_card(item_id: int, db: Session = Depends(get_db)):
    """Lấy dữ liệu Thẻ Kho (Sổ chi tiết dải vết lịch sử vật tư từ lúc nhập đến nay)"""
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Không tìm thấy vật tư")

    history = db.query(AssetHistory).filter(AssetHistory.item_id == item_id).order_by(AssetHistory.id.desc()).all()
    
    records = []
    for h in history:
        records.append({
            "id": h.id,
            "asset_code": h.asset.asset_code if h.asset else "N/A",
            "serial": h.asset.serial_number if h.asset else "N/A",
            "action_type": h.action_type,
            "performer": h.performer,
            "details": h.details,
            "timestamp": h.timestamp.isoformat()
        })

    return {
        "item_code": item.item_code,
        "name": item.name,
        "unit": item.unit,
        "current_stock": item.current_stock,
        "location": item.location,
        "history": records
    }

@router.get("/nhat-ky-nxt-excel")
def export_nxt_excel(db: Session = Depends(get_db)):
    """Trích xuất file Excel Báo Cáo Tổng Hợp Nhập - Xuất - Tồn Kho"""
    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR, exist_ok=True)

    file_path = os.path.join(EXPORTS_DIR, "bao_cao_nhap_xuat_ton.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Cáo NXT Kho"

    # Header Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "ĐOÀN NGHI LỄ CÔNG AN NHÂN DÂN - BÁO CÁO TỔNG HỢP NHẬP - XUẤT - TỒN KHO"
    ws['A1'].font = Font(name='Times New Roman', size=14, bold=True, color='991B1B')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["STT", "Mã Vật Tư", "Tên Vật Tư / Thiết Bị Kho", "Đơn Vị Tính", "Loại Kho", "Ngưỡng Cảnh Báo", "Tồn Kho Hiện Tại"]
    ws.append([]) # Blank row
    ws.append(headers)

    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    header_font = Font(name='Times New Roman', size=11, bold=True, color="FEF08A")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    items = db.query(Item).all()
    for idx, item in enumerate(items, start=1):
        row_data = [
            idx,
            item.item_code,
            item.name,
            item.unit,
            item.location or "Kho Kỹ Thuật",
            item.min_stock_alert,
            item.current_stock
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(name='Times New Roman', size=11)
            cell.border = thin_border
            if col in [1, 2, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal='center')

    wb.save(file_path)
    return FileResponse(
        path=file_path,
        filename="bao_cao_nhap_xuat_ton.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/reconciliation")
def stock_reconciliation(payload: StockReconciliationPayload, db: Session = Depends(get_db)):
    """
    Xử lý Phiếu Kiểm Kê Định Kỳ: So khớp số lượng thực tế với phần mềm, phát hiện chênh lệch Thừa / Thiếu.
    """
    scanned_codes = set(c.upper() for c in payload.scanned_asset_codes)
    all_assets = db.query(Asset).filter(Asset.status == "available").all()
    
    expected_codes = set(a.asset_code for a in all_assets)
    
    missing_codes = list(expected_codes - scanned_codes) # Thức tế không thấy (Thiếu)
    surplus_codes = list(scanned_codes - expected_codes) # Lạ hoặc chưa ghi nhận (Thừa)

    discrepancy = {
        "missing_count": len(missing_codes),
        "missing_assets": missing_codes,
        "surplus_count": len(surplus_codes),
        "surplus_assets": surplus_codes,
        "is_matched": len(missing_codes) == 0 and len(surplus_codes) == 0
    }

    import json
    sheet = AuditSheet(
        title=payload.title,
        inspector_name=payload.inspector_name,
        location=payload.location,
        scanned_count=len(scanned_codes),
        expected_count=len(expected_codes),
        discrepancy_details=json.dumps(discrepancy, ensure_ascii=False)
    )
    db.add(sheet)
    db.commit()

    log_audit(db, "storekeeper", "KIỂM KÊ KHO DỰ LỆCH", payload.title, f"Quét thực tế: {len(scanned_codes)}/{len(expected_codes)}. Thiếu: {len(missing_codes)}, Thừa: {len(surplus_codes)}")

    return {
        "success": True,
        "message": "Đã thực hiện so khớp kiểm kê thành công!",
        "audit_sheet_id": sheet.id,
        "total_scanned": len(scanned_codes),
        "total_expected": len(expected_codes),
        "discrepancy": discrepancy
    }
